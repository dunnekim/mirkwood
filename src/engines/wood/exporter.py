"""
Excel Exporter - WOOD V2 Engine

[Architecture]
Professional 9-sheet Excel model with Big 4 styling:
1. Summary
2. Assumptions  
3. WACC
4-6. DCF_Base/Bull/Bear
7. Sensitivity
8. Peer Group
9. Data Source

[Styling]
- Blue font = Input values
- Black font = Calculated values
- Professional borders and formatting
- Source attribution
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List
from datetime import datetime

from .models import ValuationSummary, Assumptions


class ExcelExporter:
    """
    Professional Excel exporter for WOOD V2 Engine
    
    [Big 4 Standard]
    - Color coding (Blue=Input, Black=Formula)
    - Professional formatting
    - Data source attribution
    - Audit trail
    """
    
    def __init__(self):
        # Define styles
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        
        self.input_font = Font(bold=False, size=10, color="0000FF")
        self.calc_font = Font(bold=False, size=10, color="000000")
        self.result_font = Font(bold=True, size=10, color="000000")
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.source_font = Font(bold=True, size=9, color="808080", italic=True)
    
    def export(
        self,
        summary: ValuationSummary,
        assumptions: Assumptions,
        peers: List[dict],
        output_dir: str
    ) -> str:
        """
        Export valuation to Excel
        
        Args:
            summary: Valuation summary
            assumptions: Assumptions
            peers: Peer list
            output_dir: Output directory
        
        Returns:
            File path
        """
        filename = f"{summary.project_name}_DCF_WOOD_V2.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Build sheets
        sheets = {}
        
        # Sheet 1: Summary
        sheets['Summary'] = self._build_summary_sheet(summary)
        
        # Sheet 2: Assumptions
        sheets['Assumptions'] = self._build_assumptions_sheet(assumptions)
        
        # Sheet 3: WACC
        sheets['WACC'] = self._build_wacc_sheet(summary)
        
        # Sheets 4-6: DCF scenarios
        for scenario in summary.scenarios:
            sheet_name = f"DCF_{scenario.scenario_name}"
            sheets[sheet_name] = self._build_dcf_sheet(scenario, assumptions)
        
        # Sheet 7: Sensitivity (Base scenario)
        sheets['Sensitivity'] = self._build_sensitivity_sheet(summary.scenarios[0], summary.base_revenue)
        
        # Sheet 8: Peer Group
        sheets['Peer_Group'] = self._build_peer_sheet(peers)
        
        # Sheet 9: Data Source
        sheets['Data_Source'] = self._build_source_sheet(summary)
        
        # Write to Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply formatting
        self._apply_formatting(filepath, summary.data_source)
        
        return filepath
    
    def _build_summary_sheet(self, summary: ValuationSummary) -> pd.DataFrame:
        """Build summary sheet"""
        data = []
        
        for scenario in summary.scenarios:
            wacc = scenario.wacc_output
            data.append({
                'Scenario': scenario.scenario_name,
                'Enterprise_Value': scenario.enterprise_value,
                'PV_FCF': scenario.pv_fcf,
                'PV_Terminal': scenario.pv_terminal,
                'WACC': wacc.wacc,
                'Cost_of_Equity': wacc.cost_of_equity,
                'Beta': wacc.beta_levered,
                'SRP': wacc.size_risk_premium
            })
        
        df = pd.DataFrame(data)
        
        # Add range summary
        range_row = {
            'Scenario': 'RANGE',
            'Enterprise_Value': f"{summary.ev_min:.0f} - {summary.ev_max:.0f}",
            'PV_FCF': '-',
            'PV_Terminal': '-',
            'WACC': '-',
            'Cost_of_Equity': '-',
            'Beta': '-',
            'SRP': '-'
        }
        
        df = pd.concat([df, pd.DataFrame([range_row])], ignore_index=True)
        
        return df
    
    def _build_assumptions_sheet(self, assumptions: Assumptions) -> pd.DataFrame:
        """Build assumptions sheet"""
        data = {
            'Parameter': [
                'Risk-Free Rate',
                'Market Risk Premium',
                'Tax Rate',
                'Terminal Growth Rate',
                'Target D/E Ratio',
                'Cost of Debt (Pre-tax)',
                'Is Listed',
                'Size Metric (Mil KRW)',
                'Projection Years'
            ],
            'Value': [
                assumptions.risk_free_rate,
                assumptions.market_risk_premium,
                assumptions.tax_rate,
                assumptions.terminal_growth_rate,
                assumptions.target_debt_ratio,
                assumptions.cost_of_debt_pretax,
                "Yes" if assumptions.is_listed else "No",
                assumptions.size_metric_mil_krw,
                assumptions.projection_years
            ]
        }
        
        return pd.DataFrame(data)
    
    def _build_wacc_sheet(self, summary: ValuationSummary) -> pd.DataFrame:
        """Build WACC detail sheet"""
        data = []
        
        for scenario in summary.scenarios:
            wacc = scenario.wacc_output
            data.append({
                'Scenario': scenario.scenario_name,
                'Risk_Free_Rate': wacc.risk_free_rate,
                'Beta_Levered': wacc.beta_levered,
                'Beta_Unlevered': wacc.beta_unlevered,
                'MRP': wacc.market_risk_premium,
                'SRP': wacc.size_risk_premium,
                'SRP_Quintile': wacc.srp_quintile,
                'Cost_of_Equity': wacc.cost_of_equity,
                'Cost_of_Debt_AT': wacc.cost_of_debt_after_tax,
                'Equity_Weight': wacc.equity_weight,
                'Debt_Weight': wacc.debt_weight,
                'WACC': wacc.wacc,
                'Beta_Source': wacc.beta_source
            })
        
        return pd.DataFrame(data)
    
    def _build_dcf_sheet(self, scenario, assumptions: Assumptions) -> pd.DataFrame:
        """Build DCF projection sheet"""
        # This is a simplified version - in production you'd reconstruct the full FCF table
        years = list(range(2026, 2026 + assumptions.projection_years))
        
        data = {
            'Year': years,
            'Revenue': scenario.revenue_projection,
            'EBITDA': scenario.ebitda_projection,
            'FCF': scenario.fcf_projection
        }
        
        df = pd.DataFrame(data)
        
        # Add summary rows
        summary_data = {
            'Year': ['', 'PV of FCF', 'PV of Terminal', 'Enterprise Value'],
            'Revenue': ['', '-', '-', '-'],
            'EBITDA': ['', '-', scenario.terminal_value_output.tv_gordon / 8.0, '-'],
            'FCF': ['', scenario.pv_fcf, scenario.pv_terminal, scenario.enterprise_value]
        }
        
        df = pd.concat([df, pd.DataFrame(summary_data)], ignore_index=True)
        
        return df
    
    def _build_sensitivity_sheet(self, base_scenario, base_revenue: float) -> pd.DataFrame:
        """Build sensitivity analysis sheet"""
        # Simplified sensitivity table
        wacc_range = [base_scenario.wacc_output.wacc + delta for delta in [-0.01, -0.005, 0, 0.005, 0.01]]
        tg_range = [0.015, 0.0175, 0.02, 0.0225, 0.025]
        
        data = []
        
        for tg in tg_range:
            row = {'Terminal_Growth': f"{tg*100:.2f}%"}
            for wacc in wacc_range:
                # Simplified EV calculation
                ev = base_scenario.enterprise_value * (1 + (base_scenario.wacc_output.wacc - wacc) * 5)
                row[f"WACC_{wacc*100:.2f}%"] = round(ev, 1)
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _build_peer_sheet(self, peers: List[dict]) -> pd.DataFrame:
        """Build peer group sheet"""
        data = []
        
        for peer in peers:
            data.append({
                'Name': peer['name'],
                'Ticker': peer.get('ticker', 'N/A'),
                'Static_Beta': peer['static_beta'],
                'Debt_Equity_Ratio': peer['debt_equity_ratio'],
                'Tax_Rate': peer['tax_rate']
            })
        
        return pd.DataFrame(data)
    
    def _build_source_sheet(self, summary: ValuationSummary) -> pd.DataFrame:
        """Build data source attribution sheet"""
        data = {
            'Item': [
                'Project Name',
                'Data Source',
                'Base Revenue (억 원)',
                'Created At',
                'Engine Version',
                'Beta Calculation',
                'Standards Applied'
            ],
            'Value': [
                summary.project_name,
                summary.data_source,
                f"{summary.base_revenue:.1f}",
                summary.created_at,
                'WOOD V2 (Nexflex Standard)',
                'Live Market Data (yfinance)',
                'KICPA (Korean IB Standard)'
            ]
        }
        
        return pd.DataFrame(data)
    
    def _apply_formatting(self, filepath: str, data_source: str):
        """
        Apply Big 4 styling to Excel file
        """
        wb = load_workbook(filepath)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Source attribution
            if ws.max_column and ws.max_column > 0:
                last_col = get_column_letter(ws.max_column)
                source_cell = ws[f'{last_col}1']
                source_cell.value = f"Source: {data_source}"
                source_cell.font = self.source_font
                source_cell.alignment = Alignment(horizontal="right", vertical="top")
            
            # Header row
            for cell in ws[1]:
                if cell.value and not str(cell.value).startswith("Source:"):
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = self.thin_border
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Data rows
            assumption_keywords = ['wacc', 'growth', 'rate', 'margin', 'ratio', 'tax', 'beta']
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    col_name = ws.cell(1, col_idx).value
                    
                    if cell.value is None:
                        continue
                    
                    cell.border = self.thin_border
                    
                    # Number formatting
                    if isinstance(cell.value, (int, float)):
                        is_assumption = False
                        if col_name:
                            col_name_lower = str(col_name).lower()
                            is_assumption = any(kw in col_name_lower for kw in assumption_keywords)
                        
                        # Percentage formatting
                        if col_name and any(x in str(col_name).lower() for x in ['wacc', 'growth', 'margin', 'rate', '%', 'weight']):
                            cell.number_format = '0.00%'
                            cell.font = self.input_font
                        else:
                            cell.number_format = '#,##0.0'
                            cell.font = self.input_font if is_assumption else self.calc_font
                        
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Assumptions sheet: all blue
            if "Assumption" in sheet_name:
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.font = self.input_font
        
        wb.save(filepath)
