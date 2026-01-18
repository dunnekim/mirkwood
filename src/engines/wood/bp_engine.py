"""
BP Engine - Business Plan Projection Engine

[Inspired by BP Engine HTML]
Porting key features to Python:
1. Historical data parsing with row exclusion
2. Driver-based projection (SaaS, Marketplace, Services, Product)
3. Excel export with formulas
4. Calculation transparency

[Difference from DCF]
- DCF: Values existing companies
- BP Engine: Projects new ventures or growth companies
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class BPEngine:
    """
    Business Plan Projection Engine
    
    Supports multiple business models:
    - SaaS/Subscription
    - Marketplace/Transaction
    - Professional Services
    - Product/Commerce
    """
    
    def __init__(self, projection_years: int = 5):
        self.projection_years = projection_years
        self.historical_data = None
        self.projections = None
    
    # ========================================================================
    # HISTORICAL DATA PROCESSING
    # ========================================================================
    
    def load_historical(
        self, 
        df: pd.DataFrame,
        exclude_keywords: List[str] = None
    ) -> Dict:
        """
        Load and clean historical data
        
        Args:
            df: DataFrame with historical P&L
                Format: Index = Accounts, Columns = Years
            exclude_keywords: Keywords to exclude (e.g., ["합계", "총계", "소계"])
        
        Returns:
            Cleaned historical data dict
        """
        if exclude_keywords is None:
            exclude_keywords = [
                "합계", "총계", "소계", "subtotal", "total", 
                "영업이익", "당기순이익", "gross profit", "operating profit",
                "EBITDA", "EBIT", "net income"
            ]
        
        # Auto-clean rows
        cleaned_df = df.copy()
        
        # Remove rows with exclude keywords
        mask = cleaned_df.iloc[:, 0].astype(str).str.contains(
            '|'.join(exclude_keywords), 
            case=False, 
            regex=True,
            na=False
        )
        
        excluded_rows = cleaned_df[mask]
        cleaned_df = cleaned_df[~mask]
        
        # Remove rows with all zeros or NaN
        numeric_cols = cleaned_df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            cleaned_df = cleaned_df[
                (cleaned_df[numeric_cols] != 0).any(axis=1) | 
                cleaned_df[numeric_cols].notna().any(axis=1)
            ]
        
        self.historical_data = {
            'cleaned': cleaned_df,
            'excluded': excluded_rows,
            'original': df
        }
        
        print(f"   📊 Historical Data Loaded:")
        print(f"      Active Rows: {len(cleaned_df)}")
        print(f"      Excluded Rows: {len(excluded_rows)}")
        
        return self.historical_data
    
    def calculate_historical_metrics(self) -> Dict:
        """
        Calculate metrics from historical data
        
        Returns:
            {
                'revenue_growth': List[float],
                'margin_trend': List[float],
                'avg_growth': float,
                'latest_margin': float
            }
        """
        if not self.historical_data:
            return {}
        
        df = self.historical_data['cleaned']
        
        # Identify revenue row (simple heuristic)
        revenue_keywords = ['매출', 'revenue', '영업수익', '수익']
        revenue_row = None
        
        for idx in df.index:
            if any(kw in str(df.iloc[idx, 0]).lower() for kw in revenue_keywords):
                revenue_row = df.loc[idx]
                break
        
        if revenue_row is None:
            return {}
        
        # Get numeric columns only
        numeric_vals = pd.to_numeric(revenue_row[1:], errors='coerce').dropna()
        
        if len(numeric_vals) < 2:
            return {}
        
        # Calculate growth rates
        growth_rates = []
        for i in range(1, len(numeric_vals)):
            if numeric_vals.iloc[i-1] > 0:
                growth = (numeric_vals.iloc[i] / numeric_vals.iloc[i-1]) - 1
                growth_rates.append(growth)
        
        avg_growth = np.mean(growth_rates) if growth_rates else 0
        
        return {
            'revenue_growth': growth_rates,
            'avg_growth': avg_growth,
            'latest_revenue': numeric_vals.iloc[-1] if len(numeric_vals) > 0 else 0,
            'years': len(numeric_vals)
        }
    
    # ========================================================================
    # PROJECTION ENGINE
    # ========================================================================
    
    def project_subscription_model(
        self,
        opening_customers: float,
        arpu: float,
        annual_churn_pct: float,
        annual_new_customers: List[float],
        annual_expansion_pct: float = 0
    ) -> pd.DataFrame:
        """
        SaaS/Subscription model projection
        
        Logic:
        - Opening customers
        - - Churned (annual_churn_pct)
        - + New customers
        - = Ending customers
        - Revenue = Avg customers × ARPU × (1 + expansion)^t
        
        Args:
            opening_customers: Starting customer count
            arpu: Average revenue per user (per year)
            annual_churn_pct: Annual customer churn (0-100)
            annual_new_customers: New customers per year [Y1, Y2, ...]
            annual_expansion_pct: ARPU expansion rate (0-100)
        
        Returns:
            DataFrame with projection
        """
        years = self.projection_years
        churn_rate = annual_churn_pct / 100
        expansion_rate = annual_expansion_pct / 100
        
        opening = [0] * years
        churned = [0] * years
        new_cust = annual_new_customers[:years] + [0] * (years - len(annual_new_customers))
        ending = [0] * years
        avg_cust = [0] * years
        arpu_per_year = [0] * years
        revenue = [0] * years
        
        for y in range(years):
            opening[y] = opening_customers if y == 0 else ending[y-1]
            churned[y] = opening[y] * churn_rate
            ending[y] = opening[y] - churned[y] + new_cust[y]
            avg_cust[y] = (opening[y] + ending[y]) / 2
            arpu_per_year[y] = arpu * ((1 + expansion_rate) ** y)
            revenue[y] = avg_cust[y] * arpu_per_year[y]
        
        df = pd.DataFrame({
            'Year': [f'Y{y+1}' for y in range(years)],
            'Opening_Customers': opening,
            'Churned': churned,
            'New_Customers': new_cust,
            'Ending_Customers': ending,
            'Avg_Customers': avg_cust,
            'ARPU': arpu_per_year,
            'Revenue': revenue
        })
        
        return df
    
    def project_with_drivers(
        self,
        model: str,
        drivers: Dict,
        cost_ratios: Dict
    ) -> pd.DataFrame:
        """
        Project financials using business drivers
        
        Args:
            model: "subscription", "transaction", "services", "product"
            drivers: Model-specific drivers
            cost_ratios: Cost structure (cogs_pct, opex_pct, etc.)
        
        Returns:
            P&L projection DataFrame
        """
        years = self.projection_years
        
        # Revenue projection (model-specific)
        if model == "subscription":
            revenue_df = self.project_subscription_model(
                opening_customers=drivers.get('opening_customers', 0),
                arpu=drivers.get('arpu', 0),
                annual_churn_pct=drivers.get('churn_pct', 15),
                annual_new_customers=drivers.get('new_customers', [0]*years),
                annual_expansion_pct=drivers.get('expansion_pct', 0)
            )
            revenue = revenue_df['Revenue'].tolist()
        else:
            # Simplified for other models
            base_revenue = drivers.get('base_revenue', 100)
            growth_rate = drivers.get('growth_rate', 0.10)
            revenue = [base_revenue * ((1 + growth_rate) ** (y+1)) for y in range(years)]
        
        # Cost projection
        cogs_pct = cost_ratios.get('cogs_pct', 0) / 100
        opex_pct = cost_ratios.get('opex_pct', 0) / 100
        
        cogs = [rev * cogs_pct for rev in revenue]
        opex = [rev * opex_pct for rev in revenue]
        ebitda = [rev - cog - op for rev, cog, op in zip(revenue, cogs, opex)]
        
        df = pd.DataFrame({
            'Year': [f'Y{y+1}' for y in range(years)],
            'Revenue': revenue,
            'COGS': cogs,
            'OPEX': opex,
            'EBITDA': ebitda
        })
        
        self.projections = df
        
        return df
    
    # ========================================================================
    # EXCEL EXPORT WITH FORMULAS
    # ========================================================================
    
    def export_to_excel_with_formulas(
        self,
        filename: str,
        include_historical: bool = True
    ) -> str:
        """
        Export to Excel with actual formulas (not just values)
        
        Features:
        - Historical tab (if available)
        - Projection tab with formulas
        - Drivers/Assumptions tab
        - Calculation breakdown tab
        
        Args:
            filename: Output filename
            include_historical: Include historical data sheet
        
        Returns:
            Filepath
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Historical (if available)
        if include_historical and self.historical_data:
            ws_hist = wb.create_sheet("Historical")
            df_hist = self.historical_data['cleaned']
            
            # Write headers
            ws_hist.append(['Account'] + list(df_hist.columns[1:]))
            
            # Write data
            for idx in df_hist.index:
                row_data = [df_hist.iloc[idx, 0]] + df_hist.iloc[idx, 1:].tolist()
                ws_hist.append(row_data)
            
            # Mark as historical (blue background)
            for row in ws_hist.iter_rows(min_row=2, max_row=ws_hist.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        # Sheet 2: Projection with FORMULAS
        if self.projections is not None:
            ws_proj = wb.create_sheet("Projection")
            df = self.projections
            
            # Write headers
            headers = df.columns.tolist()
            ws_proj.append(headers)
            
            # Write data (Year column as text, others as formulas where appropriate)
            for idx in range(len(df)):
                row_num = idx + 2
                
                # Year column
                ws_proj.cell(row_num, 1, df.iloc[idx, 0])
                
                # Revenue (input value, blue)
                ws_proj.cell(row_num, 2, df.iloc[idx, 1])
                ws_proj.cell(row_num, 2).font = Font(color="0000FF")
                
                # COGS (formula)
                if 'COGS' in headers:
                    cogs_col = headers.index('COGS') + 1
                    ws_proj.cell(row_num, cogs_col).value = f"=B{row_num}*Assumptions!$B$2"
                
                # OPEX (formula)
                if 'OPEX' in headers:
                    opex_col = headers.index('OPEX') + 1
                    ws_proj.cell(row_num, opex_col).value = f"=B{row_num}*Assumptions!$B$3"
                
                # EBITDA (formula)
                if 'EBITDA' in headers:
                    ebitda_col = headers.index('EBITDA') + 1
                    ws_proj.cell(row_num, ebitda_col).value = f"=B{row_num}-C{row_num}-D{row_num}"
        
        # Sheet 3: Assumptions
        ws_assumptions = wb.create_sheet("Assumptions")
        ws_assumptions.append(['Parameter', 'Value'])
        ws_assumptions.append(['COGS % of Revenue', 0.30])
        ws_assumptions.append(['OPEX % of Revenue', 0.40])
        ws_assumptions.append(['Growth Rate', 0.10])
        ws_assumptions.append(['Terminal Growth', 0.02])
        ws_assumptions.append(['WACC', 0.085])
        
        # Format assumptions (blue)
        for row in ws_assumptions.iter_rows(min_row=2, max_row=ws_assumptions.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.font = Font(color="0000FF", bold=True)
        
        # Save
        wb.save(filename)
        print(f"   ✅ Excel with formulas exported: {filename}")
        
        return filename
    
    # ========================================================================
    # CALCULATION BREAKDOWN
    # ========================================================================
    
    def explain_calculation(self) -> str:
        """
        Generate human-readable explanation of calculations
        
        Returns:
            Markdown formatted explanation
        """
        if not self.projections:
            return "_No projections available_"
        
        df = self.projections
        
        explanation = "# 📊 Calculation Breakdown\n\n"
        explanation += "## Revenue Projection\n\n"
        
        # Historical analysis
        if self.historical_data:
            hist_metrics = self.calculate_historical_metrics()
            if hist_metrics:
                explanation += f"**Historical Average Growth:** {hist_metrics['avg_growth']*100:.1f}%\n"
                explanation += f"**Latest Historical Revenue:** {hist_metrics['latest_revenue']:.1f}억\n\n"
        
        # Projection logic
        explanation += "**Projection Method:** Driver-based\n\n"
        
        explanation += "## P&L Structure\n\n"
        explanation += "```\n"
        explanation += "Revenue (Top-line)\n"
        explanation += "  - COGS (매출원가)\n"
        explanation += "  = Gross Profit\n"
        explanation += "  - OPEX (영업비용)\n"
        explanation += "  = EBITDA\n"
        explanation += "```\n\n"
        
        # Sample calculation for Year 1
        explanation += "## Year 1 Example\n\n"
        explanation += f"**Revenue:** {df.iloc[0]['Revenue']:.1f}억\n"
        explanation += f"**COGS:** {df.iloc[0]['COGS']:.1f}억\n"
        explanation += f"**OPEX:** {df.iloc[0]['OPEX']:.1f}억\n"
        explanation += f"**EBITDA:** {df.iloc[0]['EBITDA']:.1f}억 "
        
        margin = (df.iloc[0]['EBITDA'] / df.iloc[0]['Revenue']) * 100 if df.iloc[0]['Revenue'] > 0 else 0
        explanation += f"({margin:.1f}% margin)\n\n"
        
        return explanation
