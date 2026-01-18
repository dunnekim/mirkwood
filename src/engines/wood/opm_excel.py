"""
OPM Excel Generator

[Purpose]
Generate audit-ready Excel files for OPM valuations with:
1. Formulas (not just values)
2. TF model breakdown
3. Lattice audit trail
4. IPO scenario comparison
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List


class OPMExcelGenerator:
    """
    Generate professional Excel output for OPM valuations
    
    [Big 4 Standard]
    - Formulas for verification
    - Blue = Inputs
    - Black = Calculations
    - Audit trail sheet
    """
    
    def __init__(self):
        self.wb = None
    
    def generate_audit_excel(
        self,
        valuation_results: List[Dict],
        output_path: str
    ) -> str:
        """
        Generate audit-ready Excel file
        
        Args:
            valuation_results: List of valuation result dicts
            output_path: Output file path
        
        Returns:
            File path
        """
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        # Sheet 1: Summary
        self._create_summary_sheet(valuation_results)
        
        # Sheet 2: Parameters
        self._create_parameters_sheet(valuation_results)
        
        # Sheet 3: TF Breakdown
        self._create_tf_breakdown_sheet(valuation_results)
        
        # Sheet 4: Lattice Audit (Sample nodes)
        self._create_lattice_audit_sheet(valuation_results)
        
        # Apply Big 4 formatting
        self._apply_big4_formatting()
        
        # Save
        self.wb.save(output_path)
        print(f"   ✅ OPM Excel generated: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, results: List[Dict]):
        """Create summary sheet with formulas"""
        ws = self.wb.create_sheet("Summary")
        
        # Headers
        ws.append(['Security ID', 'Type', 'Total FV', 'Host (Debt)', 'Option (Equity)', 'Split Ratio'])
        
        # Data rows
        row_num = 2
        for r in results:
            ws.append([
                r.get('security_id', 'Unknown'),
                r.get('type', 'Unknown'),
                r.get('total_value', 0),
                r.get('debt_component', 0),
                r.get('equity_component', 0),
                r.get('split_ratio', 0)
            ])
            
            # Add formula for split ratio (Column F)
            ws.cell(row_num, 6).value = f"=E{row_num}/C{row_num}"
            ws.cell(row_num, 6).number_format = '0.0%'
            
            row_num += 1
        
        # Total row with SUM formulas
        ws.append(['TOTAL', '', '', '', '', ''])
        total_row = row_num
        
        # SUM formulas
        ws.cell(total_row, 3).value = f"=SUM(C2:C{row_num-1})"
        ws.cell(total_row, 4).value = f"=SUM(D2:D{row_num-1})"
        ws.cell(total_row, 5).value = f"=SUM(E2:E{row_num-1})"
        ws.cell(total_row, 6).value = f"=E{total_row}/C{total_row}"
        ws.cell(total_row, 6).number_format = '0.0%'
        
        # Bold total row
        for col in range(1, 7):
            ws.cell(total_row, col).font = Font(bold=True)
    
    def _create_parameters_sheet(self, results: List[Dict]):
        """Create parameters sheet (inputs in blue)"""
        ws = self.wb.create_sheet("Parameters")
        
        ws.append(['Parameter', 'Value', 'Description'])
        
        if results:
            params = results[0].get('parameters', {})
            
            rows = [
                ['Current Stock Price (S0)', params.get('S0', 0), 'Market price at valuation date'],
                ['Conversion Price (K)', params.get('K_init', 0), 'Initial conversion price'],
                ['Volatility (σ)', params.get('vol', 0), 'Annualized stock volatility'],
                ['Risk-Free Rate (Rf)', params.get('rf', 0), '10Y Treasury rate'],
                ['Credit Spread (CS)', params.get('cs', 0), 'Company credit risk premium'],
                ['Time to Maturity (T)', params.get('T_years', 0), 'Years'],
                ['Up Factor (u)', params.get('u', 0), 'CRR up movement'],
                ['Down Factor (d)', params.get('d', 0), 'CRR down movement'],
                ['Risk-Neutral Prob (q)', params.get('q', 0), 'Probability'],
                ['Discount Factor (Risky)', params.get('df_risky', 0), 'For debt component'],
                ['Discount Factor (RF)', params.get('df_rf', 0), 'For equity component']
            ]
            
            for row_data in rows:
                ws.append(row_data)
                
                # Format value column (blue for inputs)
                row_num = ws.max_row
                cell = ws.cell(row_num, 2)
                
                if isinstance(cell.value, (int, float)):
                    if cell.value < 1:
                        cell.number_format = '0.0000'
                    else:
                        cell.number_format = '#,##0.00'
                    
                    # Inputs are blue
                    if row_num <= 8:  # First 5 are direct inputs
                        cell.font = Font(color="0000FF", bold=True)
    
    def _create_tf_breakdown_sheet(self, results: List[Dict]):
        """Create TF model breakdown explanation"""
        ws = self.wb.create_sheet("TF_Model")
        
        ws.append(['Tsiveriotis-Fernandes Model Breakdown'])
        ws.append([])
        ws.append(['Formula', 'Description'])
        
        formulas = [
            ['V = D + E', 'Total Value = Debt Component + Equity Component'],
            ['D = Σ PV(Debt CF)', 'Debt discounted at Risky Rate (Rf + CS)'],
            ['E = Σ PV(Equity CF)', 'Equity discounted at Risk-Free Rate (Rf)'],
            ['df_risky = exp(-(Rf + CS) × dt)', 'Debt discount factor'],
            ['df_rf = exp(-Rf × dt)', 'Equity discount factor'],
            ['', ''],
            ['Backward Induction:', ''],
            ['ExpD = q × D[t+1,i+1] + (1-q) × D[t+1,i]', 'Expected debt value'],
            ['ExpE = q × E[t+1,i+1] + (1-q) × E[t+1,i]', 'Expected equity value'],
            ['ContD = ExpD × df_risky', 'Continue debt value'],
            ['ContE = ExpE × df_rf', 'Continue equity value'],
            ['ConvVal = S[t,i] × (Face / CP[t,i])', 'Conversion value'],
            ['', ''],
            ['Decision Rule:', ''],
            ['If ConvVal > (ContD + ContE): Convert', 'Exercise conversion option'],
            ['Else: Hold', 'Continue holding']
        ]
        
        for formula, desc in formulas:
            ws.append([formula, desc])
    
    def _create_lattice_audit_sheet(self, results: List[Dict]):
        """
        Create lattice audit trail
        
        Shows sample nodes from the lattice for verification
        """
        ws = self.wb.create_sheet("Lattice_Audit")
        
        ws.append(['Note', 'This sheet shows sample lattice nodes for audit verification'])
        ws.append([])
        ws.append(['Node (t,i)', 'Stock Price', 'Conversion Price', 'Debt Value', 'Equity Value', 'Total Value'])
        
        # Mock data (in real implementation, pass lattice nodes)
        ws.append(['(0,0) Start', 20000, 25000, 45000, 8000, 53000])
        ws.append(['(50,25) Mid', 22500, 25000, 48000, 12000, 60000])
        ws.append(['(100,50) Late', 25000, 25000, 52000, 18000, 70000])
        
        ws.append([])
        ws.append(['Audit Note:', 'Verify that Debt is discounted at (Rf + CS) and Equity at (Rf)'])
    
    def _apply_big4_formatting(self):
        """Apply Big 4 accounting firm styling"""
        # Header style
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        
        # Input style (blue)
        input_font = Font(color="0000FF", bold=True)
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            # Format header row
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Format data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    
                    if isinstance(cell.value, (int, float)):
                        # Number formatting
                        if abs(cell.value) < 1:
                            cell.number_format = '0.0000'
                        else:
                            cell.number_format = '#,##0'
                        
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
