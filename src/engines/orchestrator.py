"""
WOOD DCF Engine - Investment Banking Grade Orchestrator

[Financial Logic]
This module implements institutional-grade DCF valuation with:
1. CAPM-based WACC (Unlevered Beta → Re-levering)
2. FCF Build-up (EBIT → D&A → Capex → NWC)
3. Dual Terminal Value (Gordon Growth + Exit Multiple)
4. Sensitivity Analysis (WACC × Terminal Growth)
5. Professional Excel Formatting
"""

import json
import os
import pandas as pd
import numpy as np
from typing import Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from .wood.wacc_logic import KoreanWACCCalculator


class WoodOrchestrator:
    """
    IB-Grade DCF Valuation Engine
    
    Professional DCF model following Wall Street standards:
    - Proper WACC calculation with beta re-levering
    - Detailed FCF waterfall (EBIT → NOPAT → FCF)
    - Dual terminal value methods
    - Multi-scenario analysis with sensitivity tables
    """
    
    def __init__(self):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        self.config_path = os.path.join(current_dir, 'wood', 'config.json')
        self.output_dir = os.path.join(os.path.dirname(os.path.dirname(current_dir)), 'vault', 'reports')
        self.config = self._load_config()
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Extract key settings
        self.settings = self.config['project_settings']
        self.tax_rate = self.settings['tax_rate']
        self.rf_rate = self.settings['risk_free_rate']
        self.mrp = self.settings['market_risk_premium']
        self.terminal_g = self.settings['terminal_growth_rate']
        
        # Default assumptions (can be overridden by scenario)
        self.cost_of_debt = 0.045  # 4.5% default cost of debt
        self.exit_multiple = 8.0   # 8x EBITDA exit multiple
        
        # Initialize Korean WACC Calculator with live beta option
        use_live_beta = self.config.get('use_live_beta', False)
        self.korean_wacc = KoreanWACCCalculator(
            tax_rate=self.tax_rate,
            use_live_beta=use_live_beta
        )
        
        if use_live_beta:
            print("   🔬 Live Beta Mode: Enabled (Calculating betas from market data)")

    def _load_config(self):
        with open(self.config_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    # ========================================================================
    # WACC CALCULATION (IB Standard)
    # ========================================================================
    
    def _calculate_unlevered_beta(self, levered_beta, debt_ratio, tax_rate):
        """
        Unlever Beta using Hamada formula
        
        βu = βL / [1 + (1 - Tax) × (D/E)]
        """
        unlevered_beta = levered_beta / (1 + (1 - tax_rate) * debt_ratio)
        return unlevered_beta
    
    def _calculate_relevered_beta(self, unlevered_beta, target_debt_ratio, tax_rate):
        """
        Re-lever Beta to target capital structure
        
        βL = βu × [1 + (1 - Tax) × (D/E)]
        """
        relevered_beta = unlevered_beta * (1 + (1 - tax_rate) * target_debt_ratio)
        return relevered_beta
    
    def _calculate_wacc(
        self, 
        scenario_name, 
        scenario_params,
        target_info: Dict = None
    ):
        """
        Calculate WACC using Korean Standard (KICPA)
        
        [Korean Enhancement]
        - Size Risk Premium (SRP) based on quintile table
        - Listed vs Unlisted distinction
        - KICPA recommended MRP (8.0%)
        
        Steps:
        1. Get peer group betas and unlever them
        2. Calculate average unlevered beta
        3. Re-lever to target capital structure
        4. Apply CAPM + SRP (Korean specific)
        5. Calculate WACC: Re × (E/V) + Rd × (1-Tax) × (D/V)
        
        Args:
            scenario_name: Scenario name
            scenario_params: Scenario parameters
            target_info: Optional target company info
                {"is_listed": bool, "size_mil_krw": float}
        
        Returns:
            dict with WACC components
        """
        peer_group = self.config['peer_group']
        
        # Calculate average D/E ratio from peer group for target structure
        avg_de_ratio = np.mean([peer['debt_equity_ratio'] for peer in peer_group])
        target_debt_ratio = avg_de_ratio  # Use peer average as target
        
        # Prepare peer data for Korean WACC Calculator
        peers_formatted = []
        peer_tickers = []
        
        for peer in peer_group:
            peers_formatted.append({
                'beta': peer.get('beta', 1.0),  # Fallback beta (if live fails)
                'debt_equity_ratio': peer['debt_equity_ratio'],
                'tax_rate': peer.get('tax_rate', self.tax_rate)
            })
            
            # Collect tickers for live beta calculation
            ticker = peer.get('ticker')
            peer_tickers.append(ticker if ticker else None)
        
        # Default target info (unlisted small company if not provided)
        if target_info is None:
            target_info = {
                'is_listed': False,
                'size_mil_krw': 15000  # 150억 순자산 (5분위)
            }
        
        # Use Korean WACC Calculator (with optional live beta)
        wacc_result = self.korean_wacc.calculate(
            peers=peers_formatted,
            target_debt_ratio=target_debt_ratio,
            cost_of_debt_pretax=self.cost_of_debt,
            is_listed=target_info.get('is_listed', False),
            size_metric_mil_krw=target_info.get('size_mil_krw', 15000),
            rf=self.rf_rate,
            mrp=self.mrp,
            peer_tickers=peer_tickers  # NEW: Pass tickers for live beta
        )
        
        # Apply scenario premium
        wacc = wacc_result['WACC'] + scenario_params.get('wacc_premium', 0.0)
        
        return {
            'wacc': wacc,
            'cost_of_equity': wacc_result['Ke'],
            'cost_of_debt': self.cost_of_debt,
            'after_tax_cost_of_debt': wacc_result['Kd_post_tax'],
            'target_beta': wacc_result['Beta_Levered'],
            'unlevered_beta': wacc_result['Beta_Unlevered'],
            'debt_weight': wacc_result['Weight_Debt'],
            'equity_weight': wacc_result['Weight_Equity'],
            # Korean specific
            'srp': wacc_result['SRP'],
            'srp_quintile': wacc_result['SRP_Quintile'],
            'srp_description': wacc_result['SRP_Description'],
            'korean_standard': True
        }
    
    # ========================================================================
    # FCF BUILD-UP (The Waterfall)
    # ========================================================================
    
    def _build_fcf_projection(self, base_revenue, scenario_params, projection_years=5):
        """
        Build detailed FCF projection
        
        FCF = EBIT × (1 - Tax) + D&A - Capex - Δ NWC
        
        Returns:
            DataFrame with full FCF waterfall
        """
        growth_rate = scenario_params['revenue_growth']
        ebit_margin = scenario_params['ebit_margin']
        
        # FCF parameters (from scenario)
        da_ratio = scenario_params['da_ratio']
        capex_ratio = scenario_params['capex_ratio']
        nwc_ratio = scenario_params['nwc_ratio']
        
        # Build projection
        years = [2026 + i for i in range(projection_years)]
        revenues = [base_revenue * ((1 + growth_rate) ** (i + 1)) for i in range(projection_years)]
        
        # EBIT (direct from margin)
        ebits = [rev * ebit_margin for rev in revenues]
        
        # D&A
        das = [rev * da_ratio for rev in revenues]
        
        # EBITDA = EBIT + D&A
        ebitdas = [ebit + da for ebit, da in zip(ebits, das)]
        
        # Tax
        taxes = [ebit * self.tax_rate if ebit > 0 else 0 for ebit in ebits]
        
        # NOPAT = EBIT × (1 - Tax)
        nopats = [ebit - tax for ebit, tax in zip(ebits, taxes)]
        
        # Add back D&A (non-cash)
        nopat_plus_da = [nopat + da for nopat, da in zip(nopats, das)]
        
        # Capex
        capexs = [rev * capex_ratio for rev in revenues]
        
        # Change in NWC
        nwcs = [rev * nwc_ratio for rev in revenues]
        nwc_changes = [nwcs[0]] + [nwcs[i] - nwcs[i-1] for i in range(1, len(nwcs))]
        
        # FCF = NOPAT + D&A - Capex - Δ NWC
        fcfs = [nopat_da - capex - nwc_change 
                for nopat_da, capex, nwc_change 
                in zip(nopat_plus_da, capexs, nwc_changes)]
        
        df = pd.DataFrame({
            'Year': years,
            'Revenue': revenues,
            'EBITDA': ebitdas,
            'D&A': das,
            'EBIT': ebits,
            'Tax': taxes,
            'NOPAT': nopats,
            'Add: D&A': das,
            'Less: Capex': capexs,
            'Less: Δ NWC': nwc_changes,
            'FCF': fcfs
        })
        
        return df
    
    # ========================================================================
    # TERMINAL VALUE (Dual Method)
    # ========================================================================
    
    def _calculate_terminal_value_dual(self, last_fcf, last_ebitda, wacc):
        """
        Calculate Terminal Value using TWO methods:
        1. Perpetual Growth (Gordon Growth Model) - Primary
        2. Exit Multiple (EV/EBITDA) - Reference
        
        Returns:
            dict with both terminal values
        """
        # Method 1: Gordon Growth
        if wacc > self.terminal_g:
            tv_gordon = (last_fcf * (1 + self.terminal_g)) / (wacc - self.terminal_g)
        else:
            # Fallback if invalid
            tv_gordon = last_fcf * 20  # 20x FCF multiple
        
        # Method 2: Exit Multiple
        tv_exit = last_ebitda * self.exit_multiple
        
        # Implied multiple check
        if last_ebitda > 0:
            implied_multiple = tv_gordon / last_ebitda
        else:
            implied_multiple = 0
        
        return {
            'tv_gordon': tv_gordon,
            'tv_exit_multiple': tv_exit,
            'implied_ebitda_multiple': implied_multiple,
            'terminal_growth_rate': self.terminal_g,
            'exit_multiple_assumption': self.exit_multiple
        }
    
    # ========================================================================
    # DCF VALUATION
    # ========================================================================
    
    def _calculate_dcf(
        self, 
        scenario_name, 
        scenario_params, 
        base_revenue,
        target_info: Dict = None
    ):
        """
        Execute full DCF valuation for a scenario
        
        Args:
            scenario_name: Scenario name
            scenario_params: Scenario parameters
            base_revenue: Base revenue
            target_info: Target company info for Korean WACC
        
        Returns:
            dict with all valuation components
        """
        print(f"      📊 Scenario: {scenario_name}")
        
        # 1. Calculate WACC (Korean Standard)
        wacc_result = self._calculate_wacc(scenario_name, scenario_params, target_info)
        wacc = wacc_result['wacc']
        
        # Display Korean-specific info
        if wacc_result.get('korean_standard'):
            print(f"         WACC: {wacc*100:.2f}% (Korean KICPA Standard)")
            print(f"         SRP: {wacc_result['srp']*100:.2f}% ({wacc_result['srp_description']})")
        else:
            print(f"         WACC: {wacc*100:.2f}%")
        
        # 2. Build FCF Projection
        fcf_df = self._build_fcf_projection(base_revenue, scenario_params)
        
        # 3. Discount FCFs
        fcf_series = fcf_df['FCF'].tolist()
        discount_factors = [1 / ((1 + wacc) ** (i + 1)) for i in range(len(fcf_series))]
        pv_fcfs = [fcf * df for fcf, df in zip(fcf_series, discount_factors)]
        sum_pv_fcf = sum(pv_fcfs)
        
        fcf_df['Discount_Factor'] = discount_factors
        fcf_df['PV_FCF'] = pv_fcfs
        
        # 4. Terminal Value (Dual Method)
        last_fcf = fcf_series[-1]
        last_ebitda = fcf_df['EBITDA'].iloc[-1]
        tv_result = self._calculate_terminal_value_dual(last_fcf, last_ebitda, wacc)
        
        # Use Gordon Growth as primary
        pv_terminal = tv_result['tv_gordon'] * discount_factors[-1]
        
        # 5. Enterprise Value
        enterprise_value = sum_pv_fcf + pv_terminal
        
        print(f"         EV: {enterprise_value:.1f}억")
        
        return {
            'scenario': scenario_name,
            'wacc_result': wacc_result,
            'fcf_df': fcf_df,
            'pv_fcf': sum_pv_fcf,
            'tv_result': tv_result,
            'pv_terminal': pv_terminal,
            'enterprise_value': enterprise_value
        }
    
    # ========================================================================
    # SENSITIVITY ANALYSIS
    # ========================================================================
    
    def _generate_sensitivity_table(self, base_scenario_result, base_revenue, base_scenario_params):
        """
        Generate WACC × Terminal Growth sensitivity table
        
        Returns:
            DataFrame with sensitivity matrix
        """
        # Hardcoded sensitivity ranges (professional standard: ±1%, ±0.5%)
        wacc_range = [-0.01, -0.005, 0.0, 0.005, 0.01]
        tg_range = [-0.005, -0.0025, 0.0, 0.0025, 0.005]  # Relative to terminal_g
        
        base_wacc = base_scenario_result['wacc_result']['wacc']
        fcf_series = base_scenario_result['fcf_df']['FCF'].tolist()
        last_fcf = fcf_series[-1]
        
        # Build sensitivity matrix
        results = []
        for tg_delta in tg_range:
            adj_tg = self.terminal_g + tg_delta
            row_data = {'Terminal_Growth': f"{adj_tg*100:.2f}%"}
            for wacc_delta in wacc_range:
                adj_wacc = base_wacc + wacc_delta
                
                # Recalculate with adjusted assumptions
                discount_factors = [1 / ((1 + adj_wacc) ** (i + 1)) for i in range(len(fcf_series))]
                sum_pv_fcf = sum([fcf * df for fcf, df in zip(fcf_series, discount_factors)])
                
                if adj_wacc > adj_tg:
                    tv = (last_fcf * (1 + adj_tg)) / (adj_wacc - adj_tg)
                else:
                    tv = last_fcf * 20
                
                pv_tv = tv * discount_factors[-1]
                ev = sum_pv_fcf + pv_tv
                
                col_name = f"WACC_{adj_wacc*100:.2f}%"
                row_data[col_name] = round(ev, 1)
            
            results.append(row_data)
        
        return pd.DataFrame(results)
    
    # ========================================================================
    # EXCEL EXPORT (Professional Formatting)
    # ========================================================================
    
    def _format_excel(self, filepath, data_source: str = "User Input"):
        """
        Apply Big 4 accounting firm style formatting
        
        [Big 4 Standard]
        1. Data Source Attribution (A1 cell)
        2. Color Coding: Blue = Input, Black = Formula
        3. Professional borders and alignment
        4. Number formatting with thousand separators
        
        Args:
            filepath: Excel file path
            data_source: Data source attribution string
        """
        wb = load_workbook(filepath)
        
        # ==============================================================
        # DEFINE BIG 4 STYLES
        # ==============================================================
        
        # Header style (Dark gray + white text)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Input values (BLUE font) - Assumptions/hard-coded values
        input_font = Font(bold=False, size=10, color="0000FF")
        
        # Calculated values (BLACK font) - Formulas/linked values
        calc_font = Font(bold=False, size=10, color="000000")
        
        # Result values (BLACK BOLD) - Final outputs
        result_font = Font(bold=True, size=10, color="000000")
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # Source attribution style
        source_font = Font(bold=True, size=9, color="808080", italic=True)
        
        # ==============================================================
        # APPLY FORMATTING TO EACH SHEET
        # ==============================================================
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # ----------------------------------------------------------
            # 1. DATA SOURCE ATTRIBUTION (Right side of header)
            # ----------------------------------------------------------
            if ws.max_column and ws.max_column > 0:
                last_col = get_column_letter(ws.max_column)
                source_cell = ws[f'{last_col}1']
                source_cell.value = f"Source: {data_source}"
                source_cell.font = source_font
                source_cell.alignment = Alignment(horizontal="right", vertical="top")
            
            # ----------------------------------------------------------
            # 2. HEADER ROW FORMATTING
            # ----------------------------------------------------------
            for cell in ws[1]:
                if cell.value and cell.value != f"Source: {data_source}":
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thick_border
            
            # ----------------------------------------------------------
            # 3. AUTO-ADJUST COLUMN WIDTH
            # ----------------------------------------------------------
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            max_length = max(max_length, cell_len)
                    except:
                        pass
                adjusted_width = min(max_length + 3, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # ----------------------------------------------------------
            # 4. DATA ROW FORMATTING (Color Coding + Number Format)
            # ----------------------------------------------------------
            
            # Identify assumption columns (these should be BLUE)
            assumption_keywords = ['wacc', 'growth', 'rate', 'margin', 'ratio', 
                                  'tax', 'premium', 'beta', 'debt']
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    col_name = ws.cell(1, col_idx).value
                    
                    if cell.value is None:
                        continue
                    
                    # Apply borders
                    cell.border = thin_border
                    
                    # Number formatting and color coding
                    if isinstance(cell.value, (int, float)):
                        # Determine if this is an input or calculated value
                        is_assumption = False
                        if col_name:
                            col_name_lower = str(col_name).lower()
                            is_assumption = any(kw in col_name_lower for kw in assumption_keywords)
                        
                        # Percentage columns (inputs)
                        if col_name and any(x in str(col_name).lower() for x in ['wacc', 'growth', 'margin', 'rate', '%']):
                            cell.number_format = '0.00%'
                            cell.font = input_font  # BLUE for inputs
                        
                        # Regular numbers
                        else:
                            # Thousand separator + 1 decimal
                            cell.number_format = '#,##0.0'
                            
                            # Color code: Assumptions = Blue, Calculations = Black
                            if is_assumption or 'Scenario' in str(col_name):
                                cell.font = input_font  # BLUE
                            else:
                                # Check if it's a result row (contains "Total", "Value", "Enterprise")
                                row_label = ws.cell(row_idx, 1).value
                                if row_label and any(x in str(row_label) for x in ['Total', 'Value', 'Enterprise', 'Terminal']):
                                    cell.font = result_font  # BLACK BOLD
                                else:
                                    cell.font = calc_font  # BLACK
                        
                        # Negative numbers in red
                        if cell.value < 0:
                            current_font = cell.font
                            cell.font = Font(
                                bold=current_font.bold,
                                size=current_font.size,
                                color="FF0000"  # Red for negative
                            )
                    
                    # Text alignment
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # ----------------------------------------------------------
            # 5. SPECIAL FORMATTING FOR SPECIFIC SHEETS
            # ----------------------------------------------------------
            
            # Assumptions sheet: All values are inputs (BLUE)
            if "Assumptions" in sheet_name or "Assumption" in sheet_name:
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.font = input_font  # All BLUE
            
            # Sensitivity sheet: Highlight base case
            if "Sensitivity" in sheet_name:
                # Find and highlight base case (middle cell typically)
                if ws.max_row > 2 and ws.max_column > 2:
                    mid_row = (ws.max_row + 2) // 2
                    mid_col = (ws.max_column + 1) // 2
                    base_cell = ws.cell(mid_row, mid_col)
                    base_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    base_cell.font = Font(bold=True, size=10, color="000000")
        
        # ==============================================================
        # SAVE
        # ==============================================================
        wb.save(filepath)
        print(f"      ✨ Big 4 style formatting applied (Source: {data_source})")
    
    # ========================================================================
    # MAIN ORCHESTRATION
    # ========================================================================
    
    def run_valuation(
        self, 
        project_name: str, 
        base_revenue: float = 100.0,
        data_source: str = "User Input",
        target_info: Dict = None
    ):
        """
        Execute IB-grade DCF valuation across scenarios (Korean Standard)
        
        Args:
            project_name: Project/Company name
            base_revenue: Base year revenue (억 원)
            data_source: Data source attribution (e.g., "DART 2024.3Q")
            target_info: Target company info for Korean WACC
                {
                    "is_listed": bool,
                    "size_mil_krw": float (시가총액 or 순자산, 백만원)
                }
        
        Returns:
            (filepath, summary_text): Excel path and summary
        """
        print(f"🌲 WOOD Engine: IB-Grade DCF for '{project_name}' (Rev: {base_revenue}억)")
        print(f"   📊 Data Source: {data_source}")
        
        # Default target info (비상장 소형 가정)
        if target_info is None:
            # Revenue 기반 규모 추정 (매우 단순한 heuristic)
            estimated_size = base_revenue * 100  # 100억 매출 → 1만백만원 = 100억
            target_info = {
                'is_listed': False,
                'size_mil_krw': estimated_size
            }
            print(f"   ℹ️ Target Info: Unlisted, Estimated Size {estimated_size:,.0f}백만원")
        
        results = []
        scenarios = self.config['scenarios']
        
        # Run each scenario
        for scenario_name, scenario_params in scenarios.items():
            result = self._calculate_dcf(scenario_name, scenario_params, base_revenue, target_info)
            results.append(result)
        
        # Generate summary text
        values = [r['enterprise_value'] for r in results]
        value_range = {"min": min(values), "max": max(values), "base": results[0]['enterprise_value']}
        
        summary_text = f"🌲 **MIRKWOOD Valuation: {project_name}**\n\n"
        summary_text += f"**Enterprise Value Range: {value_range['min']:.0f}~{value_range['max']:.0f}억 원**\n"
        summary_text += f"(Base Case: {value_range['base']:.0f}억)\n\n"
        
        for r in results:
            wacc = r['wacc_result']['wacc']
            summary_text += f"**[{r['scenario']}]** EV: **{r['enterprise_value']:.1f}억** (WACC {wacc*100:.2f}%)\n"
        
        summary_text += "\n⚠️ *IB-grade DCF model with professional FCF build-up and dual terminal value.*"
        
        # Build Excel
        excel_sheets = {}
        
        # Sheet 1: Summary
        summary_data = []
        for r in results:
            wacc_res = r['wacc_result']
            summary_data.append({
                'Scenario': r['scenario'],
                'WACC': wacc_res['wacc'],
                'Cost of Equity': wacc_res['cost_of_equity'],
                'Cost of Debt (AT)': wacc_res['after_tax_cost_of_debt'],
                'Enterprise Value': r['enterprise_value'],
                'PV of FCF': r['pv_fcf'],
                'Terminal Value (PV)': r['pv_terminal']
            })
        excel_sheets['Summary'] = pd.DataFrame(summary_data)
        
        # Sheet 2: Assumptions
        base_scenario = scenarios['Base']
        assumptions_data = {
            'Parameter': [
                'Risk-Free Rate', 'Market Risk Premium', 'Cost of Debt', 
                'Tax Rate', 'Terminal Growth Rate', 'Exit Multiple (EV/EBITDA)',
                'Base: D&A Ratio', 'Base: Capex Ratio', 'Base: NWC Ratio',
                'Base: Revenue Growth', 'Base: EBIT Margin'
            ],
            'Value': [
                self.rf_rate, self.mrp, self.cost_of_debt,
                self.tax_rate, self.terminal_g, self.exit_multiple,
                base_scenario['da_ratio'],
                base_scenario['capex_ratio'],
                base_scenario['nwc_ratio'],
                base_scenario['revenue_growth'],
                base_scenario['ebit_margin']
            ]
        }
        excel_sheets['Assumptions'] = pd.DataFrame(assumptions_data)
        
        # Sheet 3-5: DCF_[Scenario]
        for r in results:
            scenario_name = r['scenario']
            df = r['fcf_df'].copy()
            
            # Add footer
            footer_rows = {
                'Year': ['Sum of PV(FCF)', 'Terminal Value (Gordon)', 
                        'Terminal Value (Exit Multiple)', 'PV of Terminal Value',
                        'Enterprise Value'],
                'Revenue': ['-', '-', '-', '-', '-'],
                'EBITDA': ['-', r['tv_result']['tv_gordon']/self.exit_multiple, 
                          r['tv_result']['tv_gordon']/self.exit_multiple, '-', '-'],
                'FCF': [r['pv_fcf'], r['fcf_df']['FCF'].iloc[-1], '-', '-', '-'],
                'PV_FCF': [r['pv_fcf'], r['tv_result']['tv_gordon'], 
                          r['tv_result']['tv_exit_multiple'], r['pv_terminal'], 
                          r['enterprise_value']]
            }
            footer_df = pd.DataFrame(footer_rows)
            
            # Combine (simplified)
            excel_sheets[f'DCF_{scenario_name}'] = df
        
        # Sheet 6: Sensitivity
        base_result = results[0]  # Use Base scenario
        base_params = scenarios['Base']
        sensitivity_df = self._generate_sensitivity_table(base_result, base_revenue, base_params)
        excel_sheets['Sensitivity'] = sensitivity_df
        
        # Export
        filename = f"{project_name}_DCF_IB_Grade.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for sheet_name, df in excel_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply Big 4 style formatting
        self._format_excel(filepath, data_source=data_source)
        
        print(f"   ✅ IB-Grade DCF Package Exported: {filepath}")
        
        return filepath, summary_text

